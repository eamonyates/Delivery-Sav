import shutil
import os
import datetime
import warnings
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side


def create_folders_and_move_files(src_directory, dest_directory, foldername, delivery_type):

        # DPX Delivery
        if delivery_type == 1: 
                shutil.move(src_directory, dest_directory)    # Move files to the destination directory
                print ('\n' + foldername + ' has been moved\n' + 'Directory moved from: ' + src_directory + '\nDirectory moved to: ' + dest_directory + '\n')    # Confirm file move to user
                os.makedirs(dest_directory + '\\DPX')    # Make DPX folder in new directory
                os.makedirs(dest_directory + '\\QT')    # Make QT folder in new directory
                print ('DPX and QT folder\'s have been created')    # Confirm folder creation to user
                
                # Find file type in the destination directory and move to correct folder, either DPX or QT
                for file in os.listdir(dest_directory):
                        
                        # Find all the QT's ending with .mov
                        if file.endswith('.mov'):
                                filepath_inc_file = dest_directory + '\\' + file    # Create the full filepath including the file itself
                                shutil.move(filepath_inc_file, dest_directory + '\\QT')    # Move the file to the QT folder
                                print ('QT moved to QT folder: ' + file)    # Inform the user that you have moved the file to the QT folder
                        
                        # Find all the DPX's
                        elif os.path.isdir(dest_directory + '\\' + file) == True:    # Check to see if the path is actually a folder
                                for item in os.listdir(dest_directory + '\\' + file):    # If it is a folder check the items inside the folder to see if the file ends with the file extension .dpx
                                        if item.endswith('.dpx'):
                                                try:
                                                        filepath_inc_file = dest_directory + '\\' + file    # Create the full filepath including the file itself
                                                        shutil.move(filepath_inc_file, dest_directory + '\\DPX')    # Move the folder containing the DPX files to the DPX folder
                                                        print ('DPX\'s moved to DPX folder: ' + file)    # Inform the user that you have moved the files to the DPX folder
                                                except shutil.Error:    
                                                        pass    # After the first DPX folder move the files will no longer exist and this will create a shutil.Error. Handle this by simply passing it without raising an error
                                        else:
                                                pass    # If file ends with anything other than .dpx then pass without action.

        # EXR Delivery
        elif delivery_type == 2: 
                shutil.move(src_directory, dest_directory)    # Move files to the destination directory
                print ('\n' + foldername + ' has been moved\n' + 'Directory moved from: ' + src_directory + '\nDirectory moved to: ' + dest_directory + '\n')    # Confirm file move to user
                os.makedirs(dest_directory + '\\EXR')    # Make EXR folder in new directory
                os.makedirs(dest_directory + '\\QT')    # Make QT folder in new directory
                print ('EXR and QT folder\'s have been created')    # Confirm folder creation to user

                # Find file type in the destination directory and move to correct folder, either EXR or QT         
                for file in os.listdir(dest_directory):
                        
                        # Find all the QT's ending with .mov
                        if file.endswith('.mov'):
                                filepath_inc_file = dest_directory + '\\' + file    # Create the full filepath including the file itself
                                shutil.move(filepath_inc_file, dest_directory + '\\QT')    # Move the file to the QT folder
                                print ('QT moved to QT folder: ' + file)    # Inform the user that you have moved the file to the QT folder
                        
                        # Find all the EXR's
                        elif os.path.isdir(dest_directory + '\\' + file) == True:    # Check to see if the path is actually a folder
                                try:
                                    for item in os.listdir(dest_directory + '\\' + file + '\\2150x1105'):    # If it is a folder check the items inside the folders folder to see if the file ends with the file extension .exr                          
                                        if item.endswith('.exr'):
                                                try:
                                                        filepath_inc_file = dest_directory + '\\' + file    # Create the full filepath including the file itself
                                                        shutil.move(filepath_inc_file, dest_directory + '\\EXR')    # Move the folder containing the EXR files to the EXR folder
                                                        print ('EXR\'s moved to EXR folder: ' + file)    # Inform the user that you have moved the files to the EXR folder
                                                except shutil.Error:    
                                                        pass    # After the first EXR folder move the files will no longer exist and this will create a shutil.Error. Handle this by simply passing it without raising an error
                                        else:
                                                pass    # If file ends with anything other than .exr then pass without action.

                                except FileNotFoundError:
                                        pass


def excel_corrections(dest_directory, foldername):

    wb = load_workbook(dest_directory + '\\' + foldername + '.xlsx')    # Get workbook
    ws = wb.worksheets[0]    # Get Worksheet
    wb_new = Workbook()    # Create new workbook to copy fields to
    ws_new = wb_new.worksheets[0]
    ws_new.title = foldername    # Change title of Worksheet

    # Find non-RV cells and copy row to new spreadsheet
    try:
        for col_cell in ws.columns[0]:

            if col_cell.coordinate == 'A1':
                empty_row_in_new_sheet = int(ws_new.max_row)        
                for row_cell in ws.rows[int(col_cell.coordinate[1:])-1]:
                    new_sheet_cell = str(row_cell.coordinate[0]) + str(empty_row_in_new_sheet)
                    ws_new[new_sheet_cell] = row_cell.value
                    ws_new.cell(row_cell.coordinate).fill = PatternFill(fill_type = 'solid', start_color = 'FF95cddc', end_color = 'FF95cddc')    # Change cell colour (The first FF is hexadecimal alpha set to 100% - http://stackoverflow.com/questions/15852122/hex-transparency-in-colors)

            elif col_cell.value[-2:] != 'rv':
                empty_row_in_new_sheet = int(ws_new.max_row) + 1

                for row_cell in ws.rows[int(col_cell.coordinate[1:])-1]:
                    new_sheet_cell = str(row_cell.coordinate[0]) + str(empty_row_in_new_sheet)
                    ws_new[new_sheet_cell] = row_cell.value
                
            else:
                pass

    except TypeError:
        pass

    # Format cells in new spreadsheet
    ws_new.row_dimensions[1].height = 16    # Set 1st rows height

    ws_new.column_dimensions["A"].width = 46.83    # Set width of column A
    ws_new.column_dimensions["B"].width = 22.16    # Set width of column B
    ws_new.column_dimensions["C"].width = 17.33    # Set width of column C
    ws_new.column_dimensions["D"].width = 63.00    # Set width of column D
    ws_new.column_dimensions["E"].width = 26.50    # Set width of column E
    ws_new.column_dimensions["F"].width = 14.16    # Set width of column F
    ws_new.column_dimensions["G"].width = 13.66    # Set width of column G
    ws_new.column_dimensions["H"].width = 13.50    # Set width of column H
    ws_new.column_dimensions["I"].width = 15.50    # Set width of column I

    i = 0
    while i < 9:
        for cell in ws_new.columns[i]:
            ws_new.cell(cell.coordinate).alignment = Alignment(wrap_text = True, horizontal = 'center', vertical = 'center')    # Set column to have center alignment and wrapped text
            ws_new.cell(cell.coordinate).border = Border(left=Side(border_style='thin', color='FF000000'), right = Side(border_style='thin', color='FF000000'), top = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'))    # Set column A to have border
            if len(cell.coordinate) == 2 and cell.coordinate[1] == '1':
                ws_new.cell(cell.coordinate).font = Font(bold = True, name = 'Arial', size = 12)    # Set Font Styles for first row
                ws_new.cell(cell.coordinate).fill = PatternFill(fill_type = 'solid', start_color = 'FF95cddc', end_color = 'FF95cddc')    # Change cell colour (The first FF is hexadecimal alpha set to 100% - http://stackoverflow.com/questions/15852122/hex-transparency-in-colors)
            else:
                ws_new.cell(cell.coordinate).font = Font(name = 'Arial', size = 11)    # Set Font Styles for remainder rows
        i += 1
          
    # Save Workbook
    wb_new.save(dest_directory + '\\' + foldername + '.xlsx')    

    # Move Workbook To Home Directory Ready for Sync
    shutil.copyfile(dest_directory + '\\' + foldername + '.xlsx', 'HOME/DIRECTORY' + foldername + '.xlsx')

    # Confirm Success
    print ('\nExcel Document Updated and Moved to HOME/DIRECTORY' + foldername + '.xlsx') 


def email_output(foldername, dest_directory, delivery_type):

    # Create Data Ops email template
    data_ops_DPX_mail_template = '''RECIPIENTS: 	
EMAIL ADDRESSES OF RECIPIENTS


SUBJECT: 
PROJECT - upload to client Server - {foldername}


MESSAGE:
Hey Data Ops,

Could you please upload the following package to the Project client Server?

Thanks,
Eamon 


-------------------------
 
FROM : 
FOLDER\TO\UPLOAD\FROM\{foldername}

TO : 
DETAILS OF UPLOAD SERVER
'''

    # Create Data Ops email template
    data_ops_EXR_mail_template = '''RECIPIENTS: 	
EMAIL ADDRESSES OF RECIPIENTS


SUBJECT: 
PROJECT - upload to Vendor Server - {foldername}


MESSAGE:
Hey Data Ops,

Can you please upload the following WIP and final EXRs to the Vendor Server for PROJECT?

Thanks,
Eamon 


-------------------------
 
FROM : 
FOLDER\TO\UPLOAD\FROM\{foldername}

TO : 
DETAILS OF UPLOAD SERVER
'''

    # Create Data Ops email template
    data_ops_sync_doc_mail_template = '''RECIPIENTS: 	
EMAIL ADDRESSES OF RECIPIENTS


SUBJECT: 
PROJECT - Sync Excel Document to Home Folder Outside sVFX


MESSAGE:
Hello Data Ops,

Could you kindly sync this excel document for me, so that I can attach it to an email going out of the Secure Network?

FILE TO SYNC: 
FILE\PATH\{workbook}

Thanks, 
Eamon
'''

    # Create client email template 1 (2 templates for beginning and end)
    client_DPX_mail_template_beginning = '''RECIPIENTS: 	
EMAIL ADDRESSES OF RECIPIENTS

SUBJECT: 
Delivery - {foldername}

MESSAGE:
Hi RECIPIENTS NAMES,

Please find the following PAFs (QTs and DPXs) uploaded to the Server for Review --

'''

    # Create client email template 2 (2 templates for beginning and end)
    client_mail_template_ending = '''
The submission form is attached.

Kind Regards,
Eamon



######################################
REMEMBER TO ATTACH THE SUBMISSION FORM
######################################
'''

    # Create client email template 1 (2 templates for beginning and end)
    client_EXR_mail_template_beginning = '''RECIPIENTS: 	
EMAIL ADDRESSES OF RECIPIENTS

SUBJECT: 
FS Delivery - {foldername}

MESSAGE:
Hi Jahanzeb and Rachael,

Please find the following Final EXRs uploaded to the Server:

'''

    # Variable for workbook name
    workbook = foldername + '.xlsx'
        
    # Create Dictionary for mail merge
    mail_merge = {'foldername':foldername, 'workbook':workbook}

    # Create Data Ops email to sync excel document in write mode
    with open('DataOps_Email_Sync_Doc.txt', 'w') as data_ops_email_sync_doc:
        data_ops_email_sync_doc.write(data_ops_sync_doc_mail_template.format(**mail_merge))

    if delivery_type == 1: 

        # Create Data Ops email in write mode
        with open('DataOps_Email_Send.txt', 'w') as data_ops_email_send:
            data_ops_email_send.write(data_ops_DPX_mail_template.format(**mail_merge))

        # Create Client email in write mode
        with open('Client_Email.txt', 'w') as client_email:
            client_email.write(client_DPX_mail_template_beginning.format(**mail_merge))

            filenames = [file for file in os.listdir(dest_directory + '\\DPX')]    # Add filenames to email
            for item in filenames:
                client_email.write("%s\n" % item)

            client_email.write(client_mail_template_ending.format(**mail_merge))

    elif delivery_type == 2:

        # Create Data Ops email in write mode
        with open('DataOps_Email_Send.txt', 'w') as data_ops_email_send:
            data_ops_email_send.write(data_ops_EXR_mail_template.format(**mail_merge))

        # Create Client email in write mode
        with open('Client_Email.txt', 'w') as client_email:
            client_email.write(client_EXR_mail_template_beginning.format(**mail_merge))

            filenames = [file for file in os.listdir(dest_directory + '\\EXR')]    # Add filenames to email
            for item in filenames:
                client_email.write("%s\n" % item)

            client_email.write(client_mail_template_ending.format(**mail_merge))

    #Confirm Success
    print ('\nEmail Templates Created\n')


# Ignore warnings (these warnings don't stop the script from running)
warnings.filterwarnings("ignore")    


#Find out what type of delivery the user is doing
print ('Please select what type of delivery you are doing:\n\n1. 2K DPX\n2. 2K EXR\n')
delivery_type = int(input('What type of delivery are you doing (select a number from the options above): '))


# Find out what delivery number is being used and add the necessary leading zeros if required
delivery_number = int(input('What delivery number are you using: '))
delivery_number_leading_zeros = "%03d" %delivery_number    # Add leading zeros


# Get today's date without hyphens
today_date = datetime.date.today()
today_date_formatted = str(today_date).replace('-','')    # Remove hyphens


# Create the directory
foldername = 'FS_' + str(delivery_number_leading_zeros) + '_2K_' + str(today_date_formatted)    # Create Folder Name
src_directory = 'SOURCE\\DIRECTORY' + foldername    # Create Destination Directory Name
dest_directory = 'DESTINATION\\DIRECTORY' + foldername    # Create Destination Directory Name


# Check to see if path exists and if it doesn't execute code
if not os.path.exists(dest_directory):
    try:
        create_folders_and_move_files(src_directory, dest_directory, foldername, delivery_type)
        excel_corrections(dest_directory, foldername)
        email_output(foldername, dest_directory, delivery_type)
		
    # Error handled if the SOURCE directory is not found
    except FileNotFoundError:
        print ('ERROR: Directory not found')    # Inform the user that the SOURCE directory can't be found

else:
    print ('ERROR: Directory already exists: ' + dest_directory)    # If directory exists, inform the user and end the process to avoid duplicate deliveries to the client

